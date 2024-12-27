from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QMessageBox, QFileDialog
)
from database import get_connection

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

class DocFinReportsForm(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Документы: Финансовая отчётность (склад)")
        self.resize(600, 400)

        layout_main = QVBoxLayout()

        self.label_invoices_sum = QLabel("Сумма всех накладных: 0.0")
        self.label_orders_sum = QLabel("Сумма всех заказов: 0.0")
        self.label_total = QLabel("Итого: 0.0")

        btn_refresh = QPushButton("Обновить данные")
        btn_export_excel = QPushButton("Вывести в Excel")

        layout_main.addWidget(self.label_invoices_sum)
        layout_main.addWidget(self.label_orders_sum)
        layout_main.addWidget(self.label_total)
        layout_main.addWidget(btn_refresh)
        layout_main.addWidget(btn_export_excel)

        self.setLayout(layout_main)

        btn_refresh.clicked.connect(self.calculate_data)
        btn_export_excel.clicked.connect(self.export_to_excel)

        self.calculate_data()

    def calculate_data(self):
        conn = get_connection()
        cursor = conn.cursor()

        # Сумма всех накладных (WarehouseInvoices.total_sum)
        cursor.execute("SELECT SUM(total_sum) FROM WarehouseInvoices")
        row = cursor.fetchone()
        invoices_sum = row[0] if row[0] is not None else 0.0

        # Сумма всех заказов (WarehouseOrders.total_sum)
        cursor.execute("SELECT SUM(total_sum) FROM WarehouseOrders")
        row = cursor.fetchone()
        orders_sum = row[0] if row[0] is not None else 0.0

        conn.close()

        total = invoices_sum + orders_sum

        self.label_invoices_sum.setText(f"Сумма всех накладных: {invoices_sum:.2f}")
        self.label_orders_sum.setText(f"Сумма всех заказов: {orders_sum:.2f}")
        self.label_total.setText(f"Итого: {total:.2f}")

    def export_to_excel(self):
        if not openpyxl:
            QMessageBox.warning(self, "Ошибка", "Модуль 'openpyxl' не установлен (pip install openpyxl).")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить фин. отчёт", "", "Excel Files (*.xlsx)")
        if not file_path:
            return

        invoices_str = self.label_invoices_sum.text()
        orders_str = self.label_orders_sum.text()
        total_str = self.label_total.text()

        wb = Workbook()
        ws = wb.active
        ws.title = "FinReport"

        bold_font = Font(bold=True)
        ws["A1"] = "Финансовая отчётность (склад)"
        ws["A1"].font = bold_font

        ws["A2"] = invoices_str
        ws["A3"] = orders_str
        ws["A4"] = total_str

        wb.save(file_path)
        QMessageBox.information(self, "Ок", f"Excel-файл '{file_path}' сохранён.")
